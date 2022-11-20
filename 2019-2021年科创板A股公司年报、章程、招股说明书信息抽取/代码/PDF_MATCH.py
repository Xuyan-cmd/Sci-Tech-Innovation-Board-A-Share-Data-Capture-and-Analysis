import json
import os

from PyPDF2 import PdfFileReader
import openpyxl
import re


# 获取pdf下面的文件夹名称
def read_PDF_DIR():
    return os.listdir("./PDF")


# 返回code列表
def read_code_excel():
    dicts = {}
    wb = openpyxl.load_workbook("./code.xlsx")
    sheet = wb.active
    rows = sheet.max_row
    for i in range(2, rows + 1):
        code = sheet.cell(i, 1).value
        name = sheet.cell(i, 2).value.replace("*", "").replace(" ", "")
        dicts[name] = code
    return dicts


# 机构投资是否含有控制权
def isControl(text):
    # 直接去看如果这个表里存在”**证券、**基金、**投资基金、**投资基金管理公司“这样子的就算控制权
    control_data = re.findall(r"前十名股东持股情况(.*?)前十名无限售条件\s?股东持股情况", text, re.S)
    if control_data:
        if "证券" in control_data[0] or "基金" in control_data[0] or "投资基金" in control_data[0] or "投资基金管理公司" in \
                control_data[0]:
            return 1
        else:
            return 0
    else:
        return 0


# 机构投资控制权一票否定权
def isNegative(text):
    # print(text)
    # print(re.findall("前十名股东持股情况",text,re.S),"="*20)
    control_data = re.findall(r"前十名股东持股情况(.*?)前十名无限售条件\s?股东持股情况", text, re.S)
    if control_data:
        lists = re.findall("\d+\.\d+", control_data[0], re.S)
        for i in lists:
            i = float(i)
            if i > 50.0:
                return 1
            else:
                return 0
    else:
        return 0


# 公司实际控制人是否通过有限合伙协议控制上市公司，是为1，否为0
def isLimited(text):
    limit_data = re.findall("前十名股东持股情况(.*?)表决权恢复的优先股股东及持股数量的说明", text, re.S)
    if limit_data:
        if re.findall("有限合伙", limit_data[0]):
            return 1
    else:
        return 0


# 匹配是否一致行动关系
def is_same_action(text):
    if re.findall("为一致行动关系", text, re.S):
        return 1
    elif re.findall("或一致行动关系", text, re.S):
        return 0
    else:
        return 0


# 获取实际控制股东情况、
def getControl(text):
    global control_person, actual_shareholderslen, actual_shareholderslen_num, same_person_flag
    actual_shareholderslen_list = []
    print(text)

    control = re.findall(r"\(一\)\s{1,}控股股东情况\s{1,}1 法人.*?(.*?)\s{1,}3", text, re.S)
    if control:
        # print("匹配成功...", control)
        control_person_match = re.findall("姓名\s{1,}(.*?)\s{1,}|单位负责人或法定代表人\s{1,}(.*?)\s{1,}", control[0])
        if not control_person_match:
            control_person = ["无"]
        else:
            control_person = ["".join(control_person_match[0]).strip()]
            # print("控股股东情况===================>", control_person)

    #  实际控制人情况
    # (二)\s{1,}实际控制人情况\s{1,}1. 法人\D+3\.
    actual = re.findall(r"\(二\)\s{1,}实际控制人情况.*?公司不存在实际控制人情况的特别说明", text, re.S)
    if actual:
        # print(actual)
        actual_person_list = re.findall("姓名\s{1,}(.*?)\s{1,}|单位负责人或法定代表人\s{1,}(.*?)\s{1,}", actual[0])

        actual_shareholderslen_num = len(actual_person_list)
        for i in actual_person_list:
            for j in i:
                if j:
                    actual_shareholderslen_list.append(j)
        # print("实际控制人情况===================>", actual_shareholderslen_list)
        same_person_flag = 1 if (control_person == actual_shareholderslen_list) else 0

    return same_person_flag, actual_shareholderslen_num


# 控股股东及其一致行动人占董事会成员比例
def getControlRate(text):
    # 董事、监事和高级管理人员的情况 董事、监事、高级管理人员和员工情况
    # print(text)
    # 现任及报告期内离任董事 、监事、 高级管理人员 和核心技术人员 持股变动及报酬情况
    # 现任及报告期内离任董事 、监事、高级管理人员 和核心技术人员 持股变动及报酬情况
    # 单位：股 姓名  职务 (注)
    # control = re.findall(r"单位：.股 姓名\s{0,}职务\s{0,}\(注\)(.*?)合计", text, re.S)
    # print(re.findall("单位：股 姓名\s+职务 \(注\)", text, re.S),"tttttttttttttttttttt")
    # control = re.findall(r"单位：股 姓名\s+职务 \(注\)(.*?)合计", text, re.S)
    # 是否为核心技术人员
    control = re.findall(r"职务\s{0,}\(注\)(.*?)合计|职务\s{0,}（注）(.*?)合计", text, re.S)
    # print(control, 2323232322)
    control = re.findall(r'是\s{0,}\s{0,}否\s{0,}在\s{0,}公\s{0,}司\s{0,}关\s{0,}联\s{0,}方\s{0,}获\s{0,}取\s{0,}报\s{0,}酬\s{0,}(.*)', str(control[0]), re.S)
    # print(control)
    person_list = re.findall('\D{2,}\s{1,}\D+\s{1,}[男|女] \d{2} \d{4}', control[0])
    print(person_list)
    ds_count = 0
    if ds_count==0:
        a = len(re.findall("独立董事",control[0]))
        b = len(re.findall("董事长", control[0]))
        c = len(re.findall("董事", control[0]))
        d = len(re.findall("董事会秘书", control[0]))

        ds_count = a+b+c+d

    gl_count = 0
    for item in person_list:
        # print(item,"sssssssssssssss")
        if "董事" in item:
            if "离任" not in item:
                # print(item)
                ds_count += 1
        if "总经理" in item or "副总经理" in item:
            if "离任" not in item:
                gl_count += 1
    if gl_count == 0:
        c = len(re.findall("总经理", str(person_list), re.S))
        c1 = len(re.findall("副总经理", str(person_list), re.S))
        gl_count = c + c1
    # print(ds_count, "董事人数")
    return ds_count, gl_count


def write_excel(data_array, na, code, sun_drop_flag):
    if not os.path.exists("./EXCEl/%s" % na):
        print("创建文件夹")
        os.mkdir("./EXCEl/%s" % na)

    wb = openpyxl.Workbook()
    print("开始写入!!!")
    sheet = wb.active
    sheet["A1"] = "stkcd"
    sheet["B1"] = "year"
    sheet["C1"] = "控股股东和实际控制人是否为同一人（是为0，否为1）"
    sheet["D1"] = "控股股东及其一致行动人占董事会成员比例"
    sheet["E1"] = "控股股东及其一致行动人占管理层成员比例"
    sheet["F1"] = "控股股东是否与其他股东有一致行动协议"
    sheet["G1"] = "公司实际控制人是否通过有限合伙协议控制上市公司，是为1，否为0"
    sheet["H1"] = "公司章程是否存在日落条款，是为1，否为0"
    sheet["I1"] = "机构投资者是否有一票否决权"
    sheet["J1"] = "机构投资者是否有控制权"
    sheet["K1"] = "备注"
    row = 2
    for item in data_array:
        a2, a3, a4, a5, a6, a7, a9, a10, a11 = item
        sheet.cell(row, 1).value = code
        sheet.cell(row, 2).value = a2
        sheet.cell(row, 3).value = a3
        sheet.cell(row, 4).value = a4
        sheet.cell(row, 5).value = a5
        sheet.cell(row, 6).value = a6
        sheet.cell(row, 7).value = a7
        sheet.cell(row, 8).value = sun_drop_flag
        sheet.cell(row, 9).value = a9
        sheet.cell(row, 10).value = a10
        sheet.cell(row, 11).value = a11
        row += 1

    wb.save("./EXCEL/%s/%s.xlsx" % (na, na))


def getTextPDF(pdfFileName, all_data):
    global num
    year = re.findall("\d{4}", pdfFileName)[0]
    pdf_file = open(pdfFileName, 'rb')
    read_pdf = PdfFileReader(pdf_file)
    text = []
    # read_pdf.getNumPages() - 1
    text = ""
    for i in range(1, read_pdf.getNumPages() - 1):
        # for i in range(1,51):
        every_page_data = read_pdf.getPage(i).extractText().replace("\n", "")
        text += every_page_data

    if re.findall(r"自然人\s{1,}√适用\s{1,}□不适用", text, re.S):
        # 控股和实际控股是否一致,返回是否一致和实际控股人数
        same_person_flag, num = getControl(text)
        # 控股股东及其一致行动人占董事会成员比例
        ds_count, gl_count = getControlRate(text)
        # 匹配是否一致行动关系
        flag = is_same_action(text)
        # 公司实际控制人是否通过有限合伙协议控制上市公司，是为1，否为0
        limit_flag = isLimited(text)
        # 机构投资是否有一票否定权
        is_Negative = isNegative(text)

        # 机构投资是否含有控制权
        is_control = isControl(text)

        if gl_count == 0:
            gl_rate = 1.0
        else:
            gl_rate = num / gl_count

        print("year", year)
        print("控股股东和实际控制人是否为同一人（是为0，否为1）", same_person_flag)
        print("控股股东及其一致行动人占董事会成员比例", num / ds_count)
        print("控股股东及其一致行动人占管理层成员比例", gl_rate)
        print("股股东是否与其他股东有一致行动协议", flag)
        print("公司实际控制人是否通过有限合伙协议控制上市公司，是为1，否为0", limit_flag)
        print("机构投资者是否有一票否决权", is_Negative)
        print("机构投资者是否有控制权", is_control)
        print("备注", " ")

        all_data.append(
            [year, same_person_flag, num / ds_count,gl_rate, flag, limit_flag, is_Negative, is_control, ""])
    else:
        print("自然人不适用")
        all_data.append(["", "", "", "", "", "", "", "", "", "自然人不适用"])


# 打开章程,返回章程的数据
def read_zhangcheng(ZCName):
    pdf_file = open(ZCName, 'rb')
    read_pdf = PdfFileReader(pdf_file)
    text = []
    ZCtext = ""
    for i in range(1, read_pdf.getNumPages() - 1):
        every_page_data = read_pdf.getPage(i).extractText().replace("\n", "")
        ZCtext += every_page_data
    return ZCtext


# 是否存在日落调控
def sundrop(ZCtext):
    if "权利应当完全相同" in ZCtext and "A 类股份" in ZCtext and "B 类股份" in ZCtext:
        print("公司章程是否存在日落条款，是为1，否为0", 1)
        return 1
    else:
        print("公司章程是否存在日落条款，是为1，否为0", 0)
        return 0


# 董事、监事和高级管理人员的情况 董事、监事、高级管理人员和员工情况
# 实际控制人情况
# 控股股东情况


code_dicts = read_code_excel()
dirs = read_PDF_DIR()
for index,d in enumerate(dirs[209:]):
    sun_drop_flag = 0
    if code_dicts.get(d):
        code = code_dicts.get(d)
        dd = "./PDF/" + d
        print(index+1, dd)
        all_data = []
        for fn in os.listdir(dd):
            filename = dd + "/" + fn
            print(filename)
            # for filename in ["PDF/吉贝尔/吉贝尔(年报2020).PDF", "PDF/吉贝尔/吉贝尔(章程).PDF"]:
            if "年报" in filename:
                try:
                    getTextPDF(filename, all_data)
                except:
                    pass

            if "章程" in filename:
                try:
                    ZCtext = read_zhangcheng(filename)
                except:
                    pass
                sun_drop_flag = sundrop(ZCtext)
        with open("./alldata.txt", "a", encoding="utf-8") as f:
            data = json.dumps([d, code, all_data, sun_drop_flag], ensure_ascii=False)
            f.write(data + "\n")
        # write_excel(all_data, d, code,sun_drop_flag)
