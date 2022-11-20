import requests
import re
import io
import os
import openpyxl
import random
from time import sleep

# 获取要下载的code
def get_download_code():
    code_lists = []
    wb = openpyxl.load_workbook("./code.xlsx")
    sheet = wb.active
    rows = sheet.max_row
    for i in range(2,rows+1):
        code = sheet.cell(i,1).value
        code_lists.append(code)
    return code_lists


# 获取pdf的函数
def get_pdf(pdf_url,company_name,key,time=""):
    response = requests.get(pdf_url)
    bytes_io = io.BytesIO(response.content)

    with open("./PDF/%s/%s(%s%s).PDF" % (company_name, company_name, key,time), mode='wb') as f:
        f.write(bytes_io.getvalue())
        print('\t\t\t\t%s(%s%s).PDF,下载成功！' % (company_name, key,time))
    # print(pdf_url)

# 返回orgID
def query_orgID():
    orgID_url = 'http://www.cninfo.com.cn/new/information/topSearch/query?keyWord=%d&maxNum=10'%keyword
    return requests.post(url=orgID_url,headers = {"User-Agent":random.choice(uas)}).json()[0]['orgId']

# 获取pdf的函数
def get_pdf_url():
    u = 'http://www.cninfo.com.cn/new/hisAnnouncement/query'
    data = {"stock": "%s,%s"%(str(keyword),str(orgID)),
            "tabName": "fulltext",
            "pageSize": "30",
            "pageNum": "1",
            "column": "sse",
            "category":" ",
            "plate": "sh",
            "seDate":"",
            "searchkey":"%s;"%key,
            "secid": "",
            "sortName": "",
            "sortType": "",
            "isHLtitle":"true"}
    data = requests.post(url=u, data=data,headers = {"User-Agent":random.choice(uas)}).json()
    if not data['announcements']:
        return
    # 若文件夹不存在则创建
    company_name = data["announcements"][0]['secName'].replace("*", "")
    if not os.path.exists("./PDF/%s" % company_name):
        print(company_name, keyword,"正在创建文件夹.........")
        os.mkdir("./PDF/%s" % company_name)

    if key =="年报":
        for item in data['announcements']:
            # year = re.findall(r'2021年<em>年度报告</em>', item["announcementTitle"])[0]
            if re.findall('.*：2021年<em>年度报告</em>$|.*2021年<em>年度报告</em>$', item["announcementTitle"]):
                pdf_url = 'http://static.cninfo.com.cn/'+item['adjunctUrl']
                get_pdf(pdf_url, company_name,key,time="2021")

            if re.findall(r'.*：2020年<em>年度报告</em>$|2020年<em>年度报告</em>$|.*2020年<em>年度报告</em>$', item["announcementTitle"]):
                pdf_url = 'http://static.cninfo.com.cn/' + item['adjunctUrl']
                get_pdf(pdf_url, company_name, key, time="2020")

            # "2019年<em>年度报告</em>"
            if re.findall(r'.*：2019年<em>年度报告</em>$|2019年<em>年度报告</em>$|.*2019年<em>年度报告</em>$', item["announcementTitle"]):
                pdf_url = 'http://static.cninfo.com.cn/' + item['adjunctUrl']
                get_pdf(pdf_url,company_name,key,time="2019")


    if key =="章程":
        pdf_url = 'http://static.cninfo.com.cn/' + data['announcements'][0]['adjunctUrl']
        get_pdf(pdf_url, company_name, key)

    if key=="招股":
        pdf_url = 'http://static.cninfo.com.cn/'+data['announcements'][0]['adjunctUrl']
        get_pdf(pdf_url,company_name,key)







if __name__ == '__main__':
    uas = ["Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.835.163 Safari/535.1",
           "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0) Gecko/20100101 Firefox/6.0",
           "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.3; .NET4.0C; .NET4.0E)",
           "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 Edg/107.0.1418.42"]
    code_lists = get_download_code()
    for index,keycode in enumerate(code_lists):
        keyword = int(keycode)
        try:
            sleep(3)
            orgID = query_orgID()
            # 招股
            keys = ["年报","招股","章程"]
            for key in keys:
                get_pdf_url()
            print("共%d个Code要下载,已经下载%d个！！！"%(len(code_lists),(index+1)))
        except:
            sleep(5)
            print("重启程序...")
            orgID = query_orgID()
            # 招股
            keys = ["年报", "招股", "章程"]
            for key in keys:
                get_pdf_url()
            print("共%d个Code要下载,已经下载%d个！！！" % (len(code_lists), (index + 1)))
