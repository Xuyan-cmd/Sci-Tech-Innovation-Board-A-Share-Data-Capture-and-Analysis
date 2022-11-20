import openpyxl
import json
import os

data2021 = []
data2020 = []
data2019 = []

def read_code():
    lists = []
    wb = openpyxl.load_workbook("./code.xlsx")
    sheet = wb.active
    rows = sheet.max_row
    for i in range(2,rows+1):
        code = sheet.cell(i,1).value
        name = sheet.cell(i,2).value
        lists.append([code,name])
    return lists
lists = read_code()

def write_excel(array):
    wb = openpyxl.Workbook()
    print("开始写入!!!")
    sheet = wb.active
    sheet["A1"] = "公司名称"
    sheet["B1"] = "stkcd"
    sheet["C1"] = "year"
    sheet["D1"] = "控股股东和实际控制人是否为同一人（是为0，否为1）"
    sheet["E1"] = "控股股东及其一致行动人占董事会成员比例"
    sheet["F1"] = "控股股东及其一致行动人占管理层成员比例"
    sheet["G1"] = "控股股东是否与其他股东有一致行动协议"
    sheet["H1"] = "公司实际控制人是否通过有限合伙协议控制上市公司，是为1，否为0"
    sheet["I1"] = "公司章程是否存在日落条款，是为1，否为0"
    sheet["J1"] = "机构投资者是否有一票否决权"
    sheet["K1"] = "机构投资者是否有控制权"
    sheet["L1"] = "备注"
    row = 2
    for item in array:
        a1, a2, a3, a4, a5, a6, a7, a8, a9,a10,a11,a12 = item
        sheet.cell(row, 1).value = a1
        sheet.cell(row, 2).value = a2
        sheet.cell(row, 3).value = a3
        sheet.cell(row, 4).value = a4
        sheet.cell(row, 5).value = a5
        sheet.cell(row, 6).value = a6
        sheet.cell(row, 7).value = a7
        sheet.cell(row, 8).value = a8
        sheet.cell(row, 9).value = a9
        sheet.cell(row, 10).value = a10
        sheet.cell(row, 11).value = a11
        sheet.cell(row, 12).value = a12
        row += 1

    wb.save("./EXCEL/%s.xlsx" % (array[0][2]))

with open("./alldata.txt", "r", encoding="utf-8") as f:
    datas = f.readlines()
    for item in datas:
        newitem = json.loads(item)
        code = newitem[1]
        company_name = newitem[0]
        if newitem[2]:
                # print(len(newitem[2]))
            for i in newitem[2]:
                # print(i)
                if i[-1] != "自然人不适用":
                    year, same_person_flag, ds_rate,gl_rate, flag, limit_flag, is_Negative, is_control,beizhu = i
                    sun_drop_flag = newitem[-1]
                    # print(company_name,code,year, same_person_flag, ds_rate,gl_rate, flag, limit_flag, is_Negative, is_control,sun_drop_flag,beizhu)
                    if year == "2021":
                        data2021.append([company_name,code,year, same_person_flag, ds_rate,gl_rate, flag, limit_flag, is_Negative, is_control,sun_drop_flag,beizhu])

                    if year == "2020":
                        data2020.append([company_name, code, year, same_person_flag, ds_rate, gl_rate, flag, limit_flag,
                                         is_Negative, is_control, sun_drop_flag, beizhu])

                    if year == "2019":
                        data2019.append([company_name, code, year, same_person_flag, ds_rate, gl_rate, flag, limit_flag,
                                         is_Negative, is_control, sun_drop_flag, beizhu])





tmp = []
for j in [data2021,data2020,data2019]:
    for i in j:
        tmp.append(i[1])



    for k in lists:
        if k[0] not in tmp:
            j.append([k[1],k[0],j[0][2],"","","","","","","","",""])

    write_excel(j)

